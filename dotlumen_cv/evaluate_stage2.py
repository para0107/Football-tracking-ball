"""
evaluate_stage2.py — Stage 2: 3D Ball Position Estimation

Runs the full pipeline (detection → CMC → Kalman → 3D estimation)
and produces:
  1. Position CSV:  3D_positions_stage2.csv
  2. Excel output:  3D_positions_stage2.xlsx   (sub-challenge 2 deliverable)
  3. Trajectory plot: trajectory_3d_stage2.png
  4. Distance over time plot: distance_stage2.png

Also supports DeepSportRadar evaluation:
    python evaluate_stage2.py --eval_deepsport --dataset_path /path/to/dataset

DeepSportRadar evaluation computes:
  - MAE depth (Z error in metres)
  - MAE full 3D position error
  - Error decomposition: detector error vs estimator error
    Using oracle radius_px from calib.compute_length2D()

Usage:
    # Run on rgb.avi (produces CSV + Excel + plots):
    python evaluate_stage2.py --video ../rgb.avi

    # Evaluate against DeepSportRadar:
    python evaluate_stage2.py --eval_deepsport \
        --dataset_path /path/to/basketball-instants-dataset

References:
    Hartley & Zisserman — Multiple View Geometry Ch. 2, 6
    Van Zandycke et al. ACM MMSports 2022 — DeepSportradar-v1
    Van Zandycke & De Vleeschouwer CVPRW 2022 — 3D ball localization
"""

import argparse
import csv
import logging
import os
import sys

import cv2
import numpy as np

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("stage2")

from config import (
    camera,
    BALL_DIAMETER_M,
    BASKETBALL_DIAMETER_M,
    ENABLE_CAMERA_COMPENSATION,
    KALMAN_MAX_MISSED_FRAMES,
    COLOR_TEXT,
)
from detector import BallDetector
from estimator import BallEstimator3D, Position3D
from motion_compensator import CameraMotionCompensator
from evaluate_stage1 import SimpleKalmanFilter


# ---------------------------------------------------------------------------
# Visualisation helpers
# ---------------------------------------------------------------------------

def draw_3d_hud(frame: np.ndarray, pos: Position3D,
                frame_idx: int) -> np.ndarray:
    """Overlay 3D position info on video frame."""
    overlay = frame.copy()
    lines = [f"Frame: {frame_idx}"]

    if pos.valid:
        lines += [
            f"X: {pos.X:+.2f} m",
            f"Y: {pos.Y:+.2f} m",
            f"Z: {pos.Z:.2f} m  (depth)",
            f"dist: {pos.distance:.2f} m",
            f"r: {pos.radius_px:.1f} px",
        ]
    else:
        lines.append("3D: no detection")

    for i, line in enumerate(lines):
        cv2.putText(overlay, line, (10, 20 + i * 22),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.55, COLOR_TEXT, 1,
                    cv2.LINE_AA)
    return overlay


def plot_trajectory_and_distance(positions: list,
                                 output_dir: str = ".") -> None:
    """
    Generate two diagnostic plots:
    1. 3D trajectory (X vs Z top-view, Y vs Z side-view)
    2. Distance from camera over time
    """
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt

        valid = [p for p in positions if p.valid]
        if len(valid) < 2:
            logger.warning("Not enough valid 3D positions to plot.")
            return

        times = [p.time_s for p in valid]
        X     = [p.X      for p in valid]
        Y     = [p.Y      for p in valid]
        Z     = [p.Z      for p in valid]
        D     = [p.distance for p in valid]

        # --- Plot 1: Trajectory (top-view XZ + side-view YZ) ---
        fig, axes = plt.subplots(1, 2, figsize=(14, 5))

        # Top view: X (lateral) vs Z (depth)
        sc = axes[0].scatter(Z, X, c=times, cmap='viridis', s=10)
        axes[0].set_xlabel("Z — Depth (m)")
        axes[0].set_ylabel("X — Lateral (m)")
        axes[0].set_title("Top-View: Ball Trajectory (X vs Z)")
        axes[0].axhline(0, color='gray', linewidth=0.5)
        axes[0].axvline(0, color='gray', linewidth=0.5)
        plt.colorbar(sc, ax=axes[0], label="Time (s)")
        axes[0].invert_xaxis()   # Z increases away from camera

        # Side view: Y (vertical, inverted for natural display) vs Z
        sc2 = axes[1].scatter(Z, [-y for y in Y], c=times, cmap='viridis', s=10)
        axes[1].set_xlabel("Z — Depth (m)")
        axes[1].set_ylabel("Height (m, up=positive)")
        axes[1].set_title("Side-View: Ball Trajectory (Y vs Z)")
        axes[1].axhline(0, color='gray', linewidth=0.5)
        plt.colorbar(sc2, ax=axes[1], label="Time (s)")
        axes[1].invert_xaxis()

        plt.tight_layout()
        path1 = os.path.join(output_dir, "trajectory_3d_stage2.png")
        plt.savefig(path1, dpi=150, bbox_inches='tight')
        plt.close()
        logger.info("Trajectory plot saved: %s", path1)

        # --- Plot 2: Distance over time ---
        fig, ax = plt.subplots(figsize=(12, 4))
        ax.plot(times, D, 'b-', linewidth=1.5, label='Distance (m)')
        ax.fill_between(times, D, alpha=0.2)
        ax.set_xlabel("Time (s)")
        ax.set_ylabel("Distance from camera (m)")
        ax.set_title("Ball Distance from Camera over Time")
        ax.legend()
        ax.grid(True, alpha=0.3)

        path2 = os.path.join(output_dir, "distance_stage2.png")
        plt.savefig(path2, dpi=150, bbox_inches='tight')
        plt.close()
        logger.info("Distance plot saved: %s", path2)

        # --- Plot 3: Sensitivity analysis ---
        # Show how depth error varies with radius_px
        r_values = np.linspace(5, 80, 200)
        estimator_tmp = BallEstimator3D(camera)
        errors = []
        for r in r_values:
            s = estimator_tmp.depth_sensitivity(r, delta_px=1.0)
            errors.append(s["depth_error_m"])

        fig, ax = plt.subplots(figsize=(10, 4))
        ax.plot(r_values, errors, 'r-', linewidth=2)
        ax.set_xlabel("radius_px (pixels)")
        ax.set_ylabel("Depth error per 1px radius error (m)")
        ax.set_title(
            "Depth Sensitivity: ΔZ = Z/radius_px\n"
            "Shows why Kalman smoothing of radius_px is critical"
        )
        ax.axvline(15, color='gray', linestyle='--', alpha=0.7,
                   label='r=15px (≈4m)')
        ax.axvline(40, color='blue', linestyle='--', alpha=0.7,
                   label='r=40px (≈1.5m)')
        ax.legend()
        ax.grid(True, alpha=0.3)

        path3 = os.path.join(output_dir, "depth_sensitivity_stage2.png")
        plt.savefig(path3, dpi=150, bbox_inches='tight')
        plt.close()
        logger.info("Sensitivity plot saved: %s", path3)

    except ImportError as e:
        logger.warning("matplotlib unavailable — skipping plots: %s", e)


def save_excel(positions: list, path: str) -> None:
    """
    Save 3D positions to Excel — sub-challenge 2 deliverable.
    Uses openpyxl for formatting.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "3D Ball Positions"

        # Header row
        headers = Position3D.csv_header()
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, pos in enumerate(positions, 2):
            for col_idx, val in enumerate(pos.to_row(), 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-width columns
        for col in ws.columns:
            max_len = max(
                len(str(cell.value or "")) for cell in col
            )
            ws.column_dimensions[
                get_column_letter(col[0].column)
            ].width = max_len + 4

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(path)
        logger.info("Excel saved: %s", path)

    except ImportError:
        logger.warning(
            "openpyxl not installed — Excel skipped. "
            "Install with: pip install openpyxl"
        )


# ---------------------------------------------------------------------------
# DeepSportRadar evaluation
# ---------------------------------------------------------------------------

def evaluate_deepsportradar(dataset_path: str,
                            estimator: BallEstimator3D) -> dict:
    """
    Evaluate 3D estimator on DeepSportRadar-v1 basketball dataset.

    Dataset provides per-image:
        item.image          — RGB numpy array
        item.calib          — full camera calibration (K, R, T)
        item.ball.center    — GT 3D position as Point3D
        item.ball.visible   — always True in ball_views.pickle

    Evaluation:
        1. Run YOLO detector on item.image → radius_px
        2. Extract fx, ppx, ppy from item.calib.K
        3. Run estimator → predicted (X, Y, Z)
        4. Compare against GT item.ball.center
        5. Compute MAE depth (Z), MAE full 3D

    Error decomposition:
        - Oracle radius_px: calib.compute_length2D(D_basketball, ball.center)
          This gives the EXACT radius_px if detection were perfect.
        - Detector error: |detected_r - oracle_r|
        - Estimator error: MAE using oracle_r (isolates formula error)
        - Combined error: MAE using detected_r (real-world performance)

    Reference:
        Van Zandycke & De Vleeschouwer CVPRW 2022 — our comparison baseline
        reports MAE on same dataset, same metric.

    Note: Ball diameter = 0.24m (basketball), not 0.22m (football).
          Adjusted in estimator call via ball_diameter_m parameter.
    """
    try:
        from mlworkflow import PickledDataset
    except ImportError:
        logger.error(
            "mlworkflow not installed. "
            "Install with: pip install mlworkflow\n"
            "Also install deepsport toolkit: "
            "pip install deepsport-toolbox"
        )
        return {}

    ball_views_path = os.path.join(dataset_path, "ball_views.pickle")
    if not os.path.exists(ball_views_path):
        logger.error(
            "ball_views.pickle not found at %s\n"
            "Run preparation script first:\n"
            "  python deepsport/scripts/prepare_ball_views_dataset.py "
            "--dataset-folder %s",
            ball_views_path, dataset_path,
        )
        return {}

    logger.info("Loading DeepSportRadar dataset from %s ...", dataset_path)
    ds = PickledDataset(ball_views_path)

    # Metrics accumulators
    errors_z_combined  = []   # depth error using detected radius
    errors_3d_combined = []   # full 3D error using detected radius
    errors_z_oracle    = []   # depth error using oracle radius
    errors_3d_oracle   = []   # full 3D error using oracle radius
    detector_r_errors  = []   # |detected_r - oracle_r|
    n_total = 0
    n_detected = 0

    detector = BallDetector()

    for key in ds.keys:
        item    = ds.query_item(key)
        image   = item.image      # numpy RGB
        calib   = item.calib
        gt      = item.ball.center

        n_total += 1

        # Convert RGB to BGR for OpenCV/YOLO
        bgr = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)

        # --- Run YOLO detector ---
        det = detector.detect(bgr)

        # --- Extract intrinsics from calib.K ---
        # calib.K is the 3x3 intrinsic matrix [[fx,0,ppx],[0,fy,ppy],[0,0,1]]
        K   = np.array(calib.K).reshape(3, 3)
        fx  = float(K[0, 0])
        fy  = float(K[1, 1])
        ppx = float(K[0, 2])
        ppy = float(K[1, 2])

        # --- Compute oracle radius_px ---
        # calib.compute_length2D(real_diameter_m, 3D_center) returns
        # the expected apparent diameter in pixels given perfect detection.
        # We use radius = diameter / 2.
        try:
            oracle_diameter_px = calib.compute_length2D(
                BASKETBALL_DIAMETER_M, gt
            )[0]
            oracle_r = oracle_diameter_px / 2.0
        except Exception:
            oracle_r = None

        # --- GT position ---
        gt_X = float(gt.x) if hasattr(gt, 'x') else float(gt[0])
        gt_Y = float(gt.y) if hasattr(gt, 'y') else float(gt[1])
        gt_Z = float(gt.z) if hasattr(gt, 'z') else float(gt[2])

        # --- Estimator using DETECTED radius ---
        if det is not None:
            n_detected += 1
            detected_r = det.radius_px
            cx, cy     = det.cx, det.cy

            # Use calib intrinsics (not D435i defaults)
            Z_pred = (fx * BASKETBALL_DIAMETER_M) / (2.0 * detected_r)
            X_pred = (cx - ppx) * Z_pred / fx
            Y_pred = (cy - ppy) * Z_pred / fy

            err_z  = abs(Z_pred - gt_Z)
            err_3d = np.sqrt(
                (X_pred-gt_X)**2 + (Y_pred-gt_Y)**2 + (Z_pred-gt_Z)**2
            )
            errors_z_combined.append(err_z)
            errors_3d_combined.append(err_3d)

            # --- Detector radius error ---
            if oracle_r is not None:
                detector_r_errors.append(abs(detected_r - oracle_r))

        # --- Estimator using ORACLE radius (isolates formula error) ---
        if oracle_r is not None and oracle_r > 0 and det is not None:
            # Use detected centre but oracle radius
            # This tells us: IF detection were perfect, how good is the formula?
            Z_oracle = (fx * BASKETBALL_DIAMETER_M) / (2.0 * oracle_r)
            X_oracle = (cx - ppx) * Z_oracle / fx
            Y_oracle = (cy - ppy) * Z_oracle / fy

            err_z_o  = abs(Z_oracle - gt_Z)
            err_3d_o = np.sqrt(
                (X_oracle-gt_X)**2 + (Y_oracle-gt_Y)**2 + (Z_oracle-gt_Z)**2
            )
            errors_z_oracle.append(err_z_o)
            errors_3d_oracle.append(err_3d_o)

        if n_total % 100 == 0:
            logger.info(
                "DeepSportRadar: processed %d items | detected %d",
                n_total, n_detected,
            )

    # --- Compute metrics ---
    metrics = {
        "n_total":          n_total,
        "n_detected":       n_detected,
        "detection_rate":   n_detected / n_total if n_total > 0 else 0,
    }

    if errors_z_combined:
        metrics.update({
            "mae_depth_combined":   float(np.mean(errors_z_combined)),
            "mae_3d_combined":      float(np.mean(errors_3d_combined)),
            "std_depth_combined":   float(np.std(errors_z_combined)),
        })
    if errors_z_oracle:
        metrics.update({
            "mae_depth_oracle":     float(np.mean(errors_z_oracle)),
            "mae_3d_oracle":        float(np.mean(errors_3d_oracle)),
            "std_depth_oracle":     float(np.std(errors_z_oracle)),
        })
    if detector_r_errors:
        metrics.update({
            "mae_radius_px":        float(np.mean(detector_r_errors)),
        })

    # --- Error decomposition ---
    if "mae_depth_combined" in metrics and "mae_depth_oracle" in metrics:
        metrics["error_from_detector_m"] = (
            metrics["mae_depth_combined"] - metrics["mae_depth_oracle"]
        )

    # --- Print summary ---
    logger.info("=" * 55)
    logger.info("DEEPSPORTRADAR EVALUATION RESULTS")
    logger.info("=" * 55)
    logger.info("Total images      : %d", metrics.get("n_total", 0))
    logger.info("Detected          : %d (%.1f%%)",
                metrics.get("n_detected", 0),
                metrics.get("detection_rate", 0) * 100)
    logger.info("")
    logger.info("--- Combined error (detected radius) ---")
    logger.info("MAE depth (Z)     : %.3f m",
                metrics.get("mae_depth_combined", float('nan')))
    logger.info("MAE full 3D       : %.3f m",
                metrics.get("mae_3d_combined", float('nan')))
    logger.info("")
    logger.info("--- Oracle error (perfect detection) ---")
    logger.info("MAE depth (Z)     : %.3f m",
                metrics.get("mae_depth_oracle", float('nan')))
    logger.info("MAE full 3D       : %.3f m",
                metrics.get("mae_3d_oracle", float('nan')))
    logger.info("")
    logger.info("--- Error decomposition ---")
    logger.info("MAE radius_px     : %.2f px",
                metrics.get("mae_radius_px", float('nan')))
    logger.info("Error from detector: %.3f m",
                metrics.get("error_from_detector_m", float('nan')))
    logger.info("Error from formula : %.3f m",
                metrics.get("mae_depth_oracle", float('nan')))
    logger.info("=" * 55)

    return metrics


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def run_stage2(video_path:   str,
               csv_path:     str = "3D_positions_stage2.csv",
               excel_path:   str = "3D_positions_stage2.xlsx",
               output_video: str = "output_stage2.avi",
               plot_dir:     str = ".") -> list:
    """
    Run Stage 2: detection → CMC → Kalman → 3D estimation on video.

    Returns list of Position3D objects (one per frame).
    """
    # --- Setup ---
    camera.load(video_path)
    logger.info("Stage 2 | %s", camera)

    detector    = BallDetector()
    estimator   = BallEstimator3D(camera)
    kf          = SimpleKalmanFilter(dt=camera.dt)
    compensator = (CameraMotionCompensator()
                   if ENABLE_CAMERA_COMPENSATION else None)

    cap = cv2.VideoCapture(video_path)
    fourcc = cv2.VideoWriter_fourcc(*'XVID')
    out    = cv2.VideoWriter(
        output_video, fourcc, camera.fps,
        (camera.frame_width, camera.frame_height)
    )

    positions         = []
    consecutive_misses = 0
    total_frames      = 0

    logger.info("Processing %d frames...", camera.total_frames)

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        total_frames += 1
        time_s        = (total_frames - 1) / camera.fps

        # --- Detection ---
        kalman_pos  = None
        allow_hough = True
        if kf.initialised:
            if consecutive_misses < KALMAN_MAX_MISSED_FRAMES:
                kalman_pos = kf.position
            elif consecutive_misses >= KALMAN_MAX_MISSED_FRAMES * 2:
                allow_hough = False

        result = detector.detect(frame, kalman_pos=kalman_pos,
                                  allow_hough=allow_hough)

        # --- Camera motion compensation ---
        if compensator is not None:
            if total_frames == 1:
                compensator.initialise(frame)
            else:
                if result is not None:
                    from detector import DetectionResult
                    cx_c, cy_c = compensator.compensate(
                        frame, result.cx, result.cy, result.radius_px
                    )
                    result = DetectionResult(
                        cx=cx_c, cy=cy_c,
                        radius_px=result.radius_px,
                        confidence=result.confidence,
                        source=result.source,
                    )
                compensator.update(
                    frame,
                    ball_cx=result.cx if result else None,
                    ball_cy=result.cy if result else None,
                    ball_radius_px=result.radius_px if result else None,
                )

        # --- Kalman ---
        if result is not None and not kf.initialised:
            kf.initialise(result.cx, result.cy, result.radius_px)

        if kf.initialised:
            kf.predict()
            if result is not None:
                kf.update([result.cx, result.cy, result.radius_px])
                consecutive_misses = 0
            else:
                consecutive_misses += 1
                if consecutive_misses >= KALMAN_MAX_MISSED_FRAMES:
                    kf.x[2, 0] = 0.0
                    kf.x[3, 0] = 0.0
                kf.x[0, 0] = float(np.clip(kf.x[0, 0], 0, camera.frame_width))
                kf.x[1, 0] = float(np.clip(kf.x[1, 0], 0, camera.frame_height))

        # --- 3D Estimation ---
        # Use Kalman-smoothed position when available,
        # fall back to raw detection, else produce null estimate.
        if kf.initialised and result is not None:
            # Use Kalman-smoothed position but detected radius_px
            # Kalman smooths cx,cy but not radius — radius is
            # detection-frame specific and used directly
            kf_cx, kf_cy = kf.position
            pos = estimator.estimate(
                frame=total_frames,
                time_s=time_s,
                cx_px=kf_cx,
                cy_px=kf_cy,
                radius_px=result.radius_px,
                source=result.source,
            )
        elif kf.initialised and consecutive_misses < KALMAN_MAX_MISSED_FRAMES:
            # Missed frame but Kalman still reliable — produce estimate
            # using smoothed position but flag as 'kalman' source
            kf_cx, kf_cy = kf.position
            # Use last known radius from positions list
            last_r = next(
                (p.radius_px for p in reversed(positions) if p.radius_px),
                None
            )
            if last_r:
                pos = estimator.estimate(
                    frame=total_frames,
                    time_s=time_s,
                    cx_px=kf_cx,
                    cy_px=kf_cy,
                    radius_px=last_r,
                    source='kalman',
                )
            else:
                pos = estimator.estimate_null(total_frames, time_s)
        else:
            pos = estimator.estimate_null(total_frames, time_s)

        positions.append(pos)

        # --- Visualise ---
        vis = draw_3d_hud(frame, pos, total_frames)
        if result is not None:
            r = int(result.radius_px)
            cx, cy = int(result.cx), int(result.cy)
            cv2.circle(vis, (cx, cy), r, (0, 255, 0), 2)
            cv2.circle(vis, (cx, cy), 3, (0, 255, 0), -1)
        out.write(vis)

        if total_frames % 50 == 0:
            valid = sum(1 for p in positions if p.valid)
            logger.info(
                "Frame %d/%d | valid 3D: %d",
                total_frames, camera.total_frames, valid,
            )

    cap.release()
    out.release()

    # --- Save CSV ---
    with open(csv_path, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(Position3D.csv_header())
        for p in positions:
            w.writerow(p.to_row())
    logger.info("CSV saved: %s", csv_path)

    # --- Save Excel ---
    save_excel(positions, excel_path)

    # --- Plots ---
    plot_trajectory_and_distance(positions, plot_dir)

    # --- Summary ---
    valid_positions = [p for p in positions if p.valid]
    if valid_positions:
        depths    = [p.Z        for p in valid_positions]
        distances = [p.distance for p in valid_positions]
        logger.info("=" * 50)
        logger.info("STAGE 2 SUMMARY")
        logger.info("=" * 50)
        logger.info("Total frames    : %d", total_frames)
        logger.info("Valid 3D frames : %d (%.1f%%)",
                    len(valid_positions),
                    len(valid_positions)/total_frames*100)
        logger.info("Depth Z range   : %.2f — %.2f m",
                    min(depths), max(depths))
        logger.info("Distance range  : %.2f — %.2f m",
                    min(distances), max(distances))
        logger.info("Avg distance    : %.2f m",
                    sum(distances)/len(distances))
        logger.info("=" * 50)

    return positions


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(
        description="Stage 2: 3D Ball Position Estimation"
    )
    p.add_argument("--video",    type=str, default=None,
                   help="Path to input video (rgb.avi)")
    p.add_argument("--csv",      type=str,
                   default="3D_positions_stage2.csv")
    p.add_argument("--excel",    type=str,
                   default="3D_positions_stage2.xlsx")
    p.add_argument("--output",   type=str,
                   default="output_stage2.avi")
    p.add_argument("--plot_dir", type=str, default=".")
    p.add_argument("--eval_deepsport", action="store_true",
                   help="Run DeepSportRadar evaluation")
    p.add_argument("--dataset_path",   type=str, default=None,
                   help="Path to basketball-instants-dataset/")
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()

    if args.video:
        run_stage2(
            video_path=args.video,
            csv_path=args.csv,
            excel_path=args.excel,
            output_video=args.output,
            plot_dir=args.plot_dir,
        )

    if args.eval_deepsport:
        if not args.dataset_path:
            logger.error(
                "--dataset_path required for DeepSportRadar evaluation.\n"
                "Download: kaggle datasets download "
                "deepsportradar/basketball-instants-dataset"
            )
            sys.exit(1)

        # Load camera for estimator (use any video or dummy config)
        if not camera.loaded and args.video:
            camera.load(args.video)
        elif not camera.loaded:
            # Dummy load for evaluation only
            import math
            camera.frame_width  = 1920
            camera.frame_height = 1080
            camera.fps          = 25.0
            camera.dt           = 1/25.0
            camera.fx = (camera.frame_width/2) / math.tan(math.radians(69)/2)
            camera.fy = camera.fx
            camera.ppx = camera.frame_width / 2
            camera.ppy = camera.frame_height / 2
            camera.total_frames = 0
            camera.loaded = True

        estimator = BallEstimator3D(camera)
        metrics   = evaluate_deepsportradar(args.dataset_path, estimator)